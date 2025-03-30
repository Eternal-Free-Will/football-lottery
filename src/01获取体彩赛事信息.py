import requests
import pandas as pd
import time
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 获取最近几期的期号列表
def get_recent_issue_list():
    url = "https://webapi.sporttery.cn/gateway/lottery/getFootBallMatchV1.qry?param=90,0&lotteryDrawNum=&sellStatus=0&termLimits=10"
    response = requests.get(url)
    data = response.json()
    return data["value"]["sfclist"]

# 获取某一期的赛事数据并生成结构分析DataFrame
def fetch_14_match_structured(lottery_draw_num):
    url = f"https://webapi.sporttery.cn/gateway/lottery/getFootBallMatchV1.qry?param=90,0&lotteryDrawNum={lottery_draw_num}&sellStatus=0&termLimits=10"
    response = requests.get(url)
    data = response.json()

    matches = data["value"]["sfcMatch"]["matchList"]
    result = []

    for m in matches:
        match_info = {
            "期号": lottery_draw_num,
            "场次": m.get("matchNum"),
            "比赛时间": m.get("startTime"),
            "联赛": m.get("matchName"),
            "主队": m.get("masterTeamAllName"),
            "客队": m.get("guestTeamAllName"),
            # 分析字段占位
            "初盘盘口": "",
            "初盘主胜赔率": "",
            "初盘平局赔率": "",
            "初盘客胜赔率": "",
            "初盘主凯利": "",
            "初盘平凯利": "",
            "初盘客凯利": "",
            "中盘盘口": "",
            "中盘主胜赔率": "",
            "中盘平局赔率": "",
            "中盘客胜赔率": "",
            "中盘主凯利": "",
            "中盘平凯利": "",
            "中盘客凯利": "",
            "临盘盘口": "",
            "临盘主胜赔率": "",
            "临盘平局赔率": "",
            "临盘客胜赔率": "",
            "临盘主凯利": "",
            "临盘平凯利": "",
            "临盘客凯利": "",
            "封盘盘口": "",
            "封盘主胜赔率": "",
            "封盘平局赔率": "",
            "封盘客胜赔率": "",
            "封盘主凯利": "",
            "封盘平凯利": "",
            "封盘客凯利": "",
            "盘口趋势观察": "",
            "凯利变化结论": "",
            "冷门信号感知（有/无）": "",
            "投注倾向（主/平/防冷/混包）": "",
            "比赛结果": ""
        }
        result.append(match_info)

    df = pd.DataFrame(result)
    df.sort_values("场次", inplace=True)
    return df

# 设置列宽
def adjust_excel_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 10
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)

# 主函数
if __name__ == "__main__":
    issues = get_recent_issue_list()
    print("获取到的期号列表：", issues)

    for issue in issues:
        df = fetch_14_match_structured(issue)

        # 构造保存路径：当前运行目录的上一级 + “足彩分析/期号”
        parent_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
        save_dir = os.path.join(parent_path, "足彩分析", issue)
        os.makedirs(save_dir, exist_ok=True)

        # 文件名和完整路径
        filename = f"传统足彩{issue}期盘口数据.xlsx"
        full_path = os.path.join(save_dir, filename)

        # 保存并调整样式
        df.to_excel(full_path, index=False)
        adjust_excel_column_width(full_path)
        print(f"已生成：{full_path}")
        time.sleep(1)
