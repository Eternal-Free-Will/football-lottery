import os
import re
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from 读取配置文件模块 import load_config

# 🎯 目标公司列表
TARGET_COMPANIES = ["36", "Bet365", "Crown", "澳门", "澳彩"]

# ✅ 从初盘页面抓取初盘赔率 & 凯利值
def get_initial_1x2_from_history(url):
    try:
        response = requests.get(url, timeout=10)
        response.encoding = "utf-8"
        soup = BeautifulSoup(response.text, "html.parser")
        rows = soup.select("table tr")

        for row in reversed(rows):
            if "(初盘)" in row.get_text():
                cols = row.find_all("td")
                if len(cols) >= 11:
                    return {
                        "初盘主胜赔率": cols[0].get_text(strip=True),
                        "初盘平局赔率": cols[1].get_text(strip=True),
                        "初盘客胜赔率": cols[2].get_text(strip=True),
                        "初盘主凯利": cols[7].get_text(strip=True),
                        "初盘平凯利": cols[8].get_text(strip=True),
                        "初盘客凯利": cols[9].get_text(strip=True)
                    }
        return None
    except Exception as e:
        print("❌ 请求初盘页面失败：", str(e))
        return None

def get_driver():
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def parse_1x2_html(html):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", id="oddsList_tab")
    if not table:
        print("❌ 没找到 oddsList_tab 表格")
        return None

    rows = table.find_all("tr")

    for row_html in rows:
        cols = row_html.find_all("td")
        if len(cols) < 12:
            continue

        company_text = cols[1].get_text(strip=True)
        if any(key in company_text for key in TARGET_COMPANIES):
            # 遍历该行所有 td，寻找 onclick 属性
            onclick_attr = ""
            for td in row_html.find_all("td"):
                onclick_raw = td.get("onclick", "")
                if "OddsHistory" in onclick_raw:
                    onclick_attr = onclick_raw
                    break

            # 然后从 onclick 中提取 id/sid/cid
            match = re.search(r"OddsHistory\('/OddsHistory\.aspx\?id=(\d+)&sid=(\d+)&cid=(\d+)", onclick_attr)
            if match:
                oid, sid, cid = match.groups()
                history_url = f"https://1x2.titan007.com/OddsHistory.aspx?id={oid}&sid={sid}&cid={cid}&l=0"
                print(f"   🎯 公司：{company_text} → 抓取初盘：{history_url}")
                result = get_initial_1x2_from_history(history_url)

                if result:
                    print("   ✅ 获取成功：", result)
                    return result
                else:
                    print("   ❌ 页面无初盘数据")
    return None

def fill_initial_1x2_odds(issue="25048"):
    parent_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    excel_path = os.path.join(parent_path, "足彩分析", issue, f"传统足彩{issue}期盘口数据补充.xlsx")

    df = pd.read_excel(excel_path, dtype=str)

    # ✅ 补字段列（如果没有）
    needed_fields = [
        "初盘主胜赔率", "初盘平局赔率", "初盘客胜赔率",
        "初盘主凯利", "初盘平凯利", "初盘客凯利"
    ]
    for field in needed_fields:
        if field not in df.columns:
            df[field] = "-"

    # ✅ 超链接提取（先加载一次原始工作簿）
    wb_orig = load_workbook(excel_path)
    ws_orig = wb_orig.active
    link_col_index = df.columns.get_loc("比赛ID") + 1

    hyperlink_map = {}
    for i in range(len(df)):
        cell = ws_orig.cell(row=i + 2, column=link_col_index)
        if cell.hyperlink:
            hyperlink_map[i] = cell.hyperlink.target

    # ✅ 抓取数据
    for i, row in df.iterrows():
        link = hyperlink_map.get(i)
        if not link or not link.startswith("http"):
            print(f"⏭️ 跳过第 {i+1} 行：无有效链接")
            continue

        match = re.search(r"id=(\d+)", link)
        if not match:
            print(f"⏭️ 跳过第 {i+1} 行：链接格式异常")
            continue

        match_id = match.group(1)
        url = f"https://1x2.titan007.com/oddslist/{match_id}.htm"
        print(f"\n➡️ 抓取 第{i+1}行 比赛ID：{match_id}")

        try:
            driver = get_driver()
            driver.set_page_load_timeout(20)
            driver.get(url)

            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            WebDriverWait(driver, 20).until(
                lambda d: len(d.find_elements(By.XPATH, '//table[@id="oddsList_tab"]/tbody/tr')) >= 5
            )

            html = driver.page_source
            result = parse_1x2_html(html)
            driver.quit()

            if result and isinstance(result, dict):
                for key, val in result.items():
                    df.at[i, key] = val
                print("✅ 写入成功：", result)
            else:
                print("❌ 页面结构异常或无目标公司")
                for field in needed_fields:
                    df.at[i, field] = "-"

        except Exception as e:
            print(f"❌ 报错 比赛ID {match_id}：{str(e)}")
            for field in needed_fields:
                df.at[i, field] = "-"
            try:
                driver.quit()
            except:
                pass

    # ✅ 写入 Excel 文件
    df.to_excel(excel_path, index=False)

    # ✅ 加载写入后的文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # ✅ 设置列宽（可自定义）
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = min(max_len + 12, 30)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # ✅ 设置表头样式（加粗 + 居中）
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    # ✅ 恢复超链接
    for i, url in hyperlink_map.items():
        cell = ws.cell(row=i + 2, column=link_col_index)
        cell.value = "查看盘口"
        cell.hyperlink = url
        cell.style = "Hyperlink"

    # ✅ 保存
    wb.save(excel_path)
    print(f"\n✅ 表格已美化并保存：{excel_path}")

    print(f"\n✅ 全部数据处理完成，已保存：{excel_path}")

# 运行入口
if __name__ == "__main__":
    issue, date_str = load_config()
    print("当前期号:", issue)
    fill_initial_1x2_odds(issue)
