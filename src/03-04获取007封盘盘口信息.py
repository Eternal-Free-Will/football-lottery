import os
import re
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from 读取配置文件模块 import load_config

# 启动浏览器
def get_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    return webdriver.Chrome(options=options)

# 模糊匹配公司名
def is_target_company(name):
    for keyword in ["36", "Crown", "澳门", "澳彩"]:
        if keyword in name:
            return True
    return False

# 盘口文本 → 数值转换函数
def convert_handicap(text):
    text = text.replace(" ", "").replace("\xa0", "")
    
    mapping = {
        "平手": 0.0, "平": 0.0,
        "平手/半球": -0.25, "半球": -0.5, "半球/一球": -0.75, "一球": -1.0,
        "一球/球半": -1.25, "球半": -1.5, "球半/两球": -1.75, "两球": -2.0,
        "两球/两球半": -2.25, "两球半": -2.5, "两球半/三球": -2.75, "三球": -3.0,
        "三球/三球半": -3.25, "三球半": -3.5, "三球半/四球": -3.75, "四球": -4.0,
        "四球/四球半": -4.25, "四球半": -4.5, "四球半/五球": -4.75, "五球": -5.0,

        "受让平手/半球": 0.25, "受让半球": 0.5, "受让半球/一球": 0.75, "受让一球": 1.0,
        "受让一球/球半": 1.25, "受让球半": 1.5, "受让球半/两球": 1.75, "受让两球": 2.0,
        "受让两球/两球半": 2.25, "受让两球半": 2.5, "受让两球半/三球": 2.75, "受让三球": 3.0,
        "受让三球/三球半": 3.25, "受让三球半": 3.5, "受让三球半/四球": 3.75, "受让四球": 4.0,
        "受让四球/四球半": 4.25, "受让四球半": 4.5, "受让四球半/五球": 4.75, "受让五球": 5.0
    }

    return mapping.get(text, None)

# 获取封盘盘口数值
def fetch_initial_handicap(driver, match_id):
    url = f"https://vip.titan007.com/AsianOdds_n.aspx?id={match_id}"
    driver.get(url)
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.ID, "odds"))
        )
    except:
        print(f"⚠️ 页面加载失败：{url}")
        return None

    rows = driver.find_elements(By.XPATH, '//table[@id="odds"]/tbody/tr')
    fallback = None

    for row in rows:
        try:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) < 5:
                continue
            company = cols[0].text.strip()
            handicap = cols[6].text.strip()
            if not handicap:
                # 封盘之后，网页对于中间三个td采用了display:none，所以需要从第9个元素中获取
                handicap = cols[9].text.strip()
            value = convert_handicap(handicap)
            if value is not None and is_target_company(company):
                return str(value)
            if value is not None and fallback is None:
                fallback = str(value)
        except:
            continue

    return fallback

# 主函数
def fill_initial_handicap(issue="25048"):
    parent_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    excel_path = os.path.join(
        parent_path, "足彩分析", issue, f"传统足彩{issue}期盘口数据补充.xlsx"
    )

    df = pd.read_excel(excel_path, dtype=str)
    wb = load_workbook(excel_path)
    ws = wb.active

    link_col_index = df.columns.get_loc("比赛ID") + 1
    hyperlink_map = {}

    for i in range(len(df)):
        cell = ws.cell(row=i + 2, column=link_col_index)
        if cell.hyperlink:
            hyperlink_map[i] = cell.hyperlink.target

    driver = get_driver()

    for i, row in df.iterrows():
        link = hyperlink_map.get(i)
        if not link or not link.startswith("http"):
            print(f"跳过第 {i+1} 行：无有效链接")
            continue

        match = re.search(r"id=(\d+)", link)
        if not match:
            print(f"跳过第 {i+1} 行：链接格式异常")
            continue

        match_id = match.group(1)
        print(f"➡️ 抓取盘口（第{i+1}行）比赛ID：{match_id}")

        value = fetch_initial_handicap(driver, match_id)
        if value is not None:
            df.at[i, "封盘盘口"] = value
            print(f"✅ 写入：封盘盘口 = {value}")
        else:
            df.at[i, "封盘盘口"] = "-"
            print(f"❌ 未获取盘口信息：{match_id}")

    driver.quit()

    # 保存数据
    df.to_excel(excel_path, index=False)

    # ✅ 加载写入后的文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # ✅ 设置列宽（可自定义）
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = min(max_len + 12, 30)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # ✅ 设置表头样式（加粗 + 居中）
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    # ✅ 恢复超链接
    for i, url in hyperlink_map.items():
        cell = ws.cell(row=i + 2, column=link_col_index)
        cell.value = "查看盘口"
        cell.hyperlink = url
        cell.style = "Hyperlink"

    # ✅ 保存
    wb.save(excel_path)
    print(f"✅ 表格已更新并保存：{excel_path}")

# 执行
if __name__ == "__main__":
    issue, date_str = load_config()
    print("当前期号:", issue)
    fill_initial_handicap(issue)
