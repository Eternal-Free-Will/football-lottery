import os
import re
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from 读取配置文件模块 import load_config

# 清洗队名
def clean_team_name(name):
    name = re.sub(r'\[.*?\]', '', str(name))
    name = re.sub(r'\s+', '', name)
    return name.strip()

# 加载队名映射表
def load_team_mapping(csv_path="球队名称映射表.csv"):
    if os.path.exists(csv_path):
        df_map = pd.read_csv(csv_path, dtype=str)
        df_map.dropna(inplace=True)
        return {
            clean_team_name(k): clean_team_name(v)
            for k, v in zip(df_map['excel_team'], df_map['titan007_team'])
        }
    return {}

# 启动 Selenium
def get_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--window-size=1920,1080')
    return webdriver.Chrome(options=options)

# 获取页面 HTML
def fetch_matches_html_by_date(date_str):
    url = f"https://bf.titan007.com/football/Next_{date_str}.htm"
    driver = get_driver()
    driver.get(url)
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "tr"))
        )
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
    except:
        print("⚠️ 页面加载等待超时或结构变化")
    html = driver.page_source
    driver.quit()
    with open("debug.html", "w", encoding="utf-8") as f:
        f.write(html)
    return html

# 解析比赛数据
def extract_matches_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    rows = soup.find_all("tr")

    data = []
    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 7:
            continue
        try:
            league = cols[1].get_text(strip=True)
            home = clean_team_name(cols[3].get_text(strip=True))
            away = clean_team_name(cols[5].get_text(strip=True))
            row_html = str(row)
            match = re.search(r"AsianOdds\((\d+)\)", row_html)
            match_id = match.group(1) if match else ""
            if home and away and match_id:
                data.append({
                    "联赛": league,
                    "主队": home,
                    "客队": away,
                    "比赛ID": match_id
                })
        except:
            continue

    if not data:
        print("❌ 页面中未成功提取任何比赛数据！")
    return pd.DataFrame(data, columns=["联赛", "主队", "客队", "比赛ID"])

# 主函数：填入比赛ID和匹配状态
def fill_excel_with_match_ids(issue="25048", date_str="20250329"):
    parent_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    input_path = os.path.join(parent_path, "足彩分析", issue, f"传统足彩{issue}期盘口数据.xlsx")
    output_path = os.path.join(parent_path, "足彩分析", issue, f"传统足彩{issue}期盘口数据补充.xlsx")

    df_excel = pd.read_excel(input_path, dtype=str)

    # 插入比赛ID与匹配状态列（在“场次”后）
    match_index = df_excel.columns.get_loc("场次")
    df_excel.insert(match_index + 1, "比赛ID", "")
    df_excel.insert(match_index + 2, "匹配状态", "")

    html = fetch_matches_html_by_date(date_str)
    df_html = extract_matches_from_html(html)
    team_map = load_team_mapping()

    for i, row in df_excel.iterrows():
        raw_home = clean_team_name(row['主队'])
        raw_away = clean_team_name(row['客队'])
        home = team_map.get(raw_home, raw_home)
        away = team_map.get(raw_away, raw_away)

        match_row = df_html[(df_html['主队'] == home) & (df_html['客队'] == away)]

        if not match_row.empty:
            match_id = match_row.iloc[0]['比赛ID']
            match_url = f"https://vip.titan007.com/AsianOdds_n.aspx?id={match_id}"
            df_excel.at[i, "比赛ID"] = match_url
            df_excel.at[i, "匹配状态"] = "成功"
            print(f"✅ 匹配：{raw_home}({home}) vs {raw_away}({away})")
        else:
            df_excel.at[i, "比赛ID"] = "-"
            df_excel.at[i, "匹配状态"] = "未匹配"
            print(f"❌ 未匹配：{raw_home}({home}) vs {raw_away}({away})")

    # 写入 Excel
    df_excel.to_excel(output_path, index=False)

    # 设置列宽和超链接
    wb = load_workbook(output_path)
    ws = wb.active
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)

        if col[0].value == "比赛ID":
            for cell in col[1:]:
                url = str(cell.value)
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value = "查看盘口"
                    cell.style = "Hyperlink"
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = max(max_length + 10, 10)

    wb.save(output_path)
    print(f"✅ 新文件已保存：{output_path}")

# 入口
if __name__ == "__main__":
    issue, date_str = load_config()
    print("当前期号:", issue)
    print("当前日期:", date_str)
    fill_excel_with_match_ids(issue, date_str)
